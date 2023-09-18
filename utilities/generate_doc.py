import openpyxl
from docx import Document

def generate_doc(table_data):

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    for row_data in table_data:
        sheet.append(row_data)
    return workbook

def convert_json_to_docx_table(json_data):
    try:
        doc = Document()
        table = doc.add_table(rows=1, cols=len(json_data[0].keys()))

        for i, col_name in enumerate(json_data[0].keys()):
            table.cell(0, i).text = col_name

        for row in json_data:
            for i, col_value in enumerate(row.values()):
                table.cell(i + 1, i).text = str(col_value)
        return doc
    except Exception as e:
        print('error converting json to table', e)
